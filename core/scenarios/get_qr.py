from __future__ import annotations

from .base import BaseScenario
from .modes import ScenarioMode
from playwright.async_api import TimeoutError as PlaywrightTimeoutError

from typing import List, Dict, Any, Tuple
import base64
from pathlib import Path
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook



class QRcodeScenario(BaseScenario):
    mode = ScenarioMode.QRCODE

    # ✅ ИЗМЕНЕНО ДЛЯ МАССОВОГО СБОРА:
    # теперь сценарий НЕ пишет Excel, а ВОЗВРАЩАЕТ данные
    async def run(self) -> Dict[str, Any]:
        c = self.c
        acc = c.account

        phone10 = acc["phone10"]
        account_name = acc.get("name", "")

        print("[SCENARIO] QR_CODE")
        c.on_progress and c.on_progress(5, "Открываю сайт…")

        await c. accept_cookie()
        await c.close_modal()
        await c.wait_full_load()
        await c.humanize()

        await self.click_btn_order()
        c.on_progress and c.on_progress(10, "Кнопка заказы нажата")

        await c.human_wait()
        await c.change_window_size(800, 1200)
        await c.human_wait()

        # --- СКАН ПВЗ + ТОВАРЫ ---
        pvz_list = await scan_pvz_with_products(c.page)

        for pvz in pvz_list:
            print("ПВЗ:", pvz.get("adress_pvz", ""))
            for p in pvz.get("products", []):
                print("  SKU:", p.get("sku", ""), "status:", p.get("status", ""))

        # --- QR ---
        try:
            c.on_progress and c.on_progress(15, "Жду QR блок…")
            code, qr_base64 = await fetch_qr_code_data(c.page)
            c.on_progress and c.on_progress(40, f"QR получен, код: {code}")
        except PlaywrightTimeoutError:
            code, qr_base64 = "", ""

        # ✅ ВАЖНО: возвращаем данные для общего сохранения ПОСЛЕ всех аккаунтов
        return {
            "phone10": phone10,
            "account_name": account_name,
            "pvz_list": pvz_list,
            "code": code,
            "qr_base64": qr_base64,
        }

    async def click_btn_order(self):
        c = self.c
        order_btn = await c.page.wait_for_selector(
            'a[data-wba-header-name="DLV"]',
            timeout=5000,
            state="visible"
        )
        await order_btn.hover()
        await c.human_wait()
        await c.human_click(order_btn)


async def fetch_qr_code_data(page) -> Tuple[str, str]:
    await page.wait_for_selector("section.delivery-qr", timeout=15000, state="visible")

    code_el = await page.wait_for_selector(".delivery-qr__main span", timeout=5000, state="visible")
    code = (await code_el.inner_text()).replace(" ", "").strip()

    img = await page.wait_for_selector(".delivery-qr__code-wrap img", timeout=5000, state="visible")
    src = await img.get_attribute("src") or ""

    qr_base64 = src.split(",", 1)[1] if "," in src else ""
    return code, qr_base64


# ---------------- SKU SPLIT ----------------

def sku_from_src(src: str) -> str:
    if not src:
        return ""
    try:
        # .../partXXXX/SKU/images/...
        return src.split("/part", 1)[1].split("/", 2)[1]
    except Exception:
        return ""


# ---------------- PVZ + PRODUCTS ----------------

async def scan_pvz_with_products(page) -> List[Dict[str, Any]]:
    await page.wait_for_selector('[data-testid="delivery-page"]', timeout=15000)

    for _ in range(6):
        await page.mouse.wheel(0, 1000)
        await page.wait_for_timeout(300)

    blocks = page.locator("div.delivery-block__content")
    n = await blocks.count()

    out: List[Dict[str, Any]] = []

    for i in range(n):
        block = blocks.nth(i)

        addr_loc = block.locator("p.delivery-address__info")
        if await addr_loc.count() == 0:
            continue

        adress = (await addr_loc.first.inner_text()).strip()
        if not adress:
            continue

        items = block.locator(".delivery-block__list .custom-slider__item.product")
        m = await items.count()

        products: List[Dict[str, Any]] = []

        for j in range(m):
            it = items.nth(j)

            status = ""
            st = it.locator("span.product__tracking")
            if await st.count() > 0:
                status = (await st.first.inner_text()).strip()

            img = it.locator(".product__photo img")
            src = await img.first.get_attribute("src") if await img.count() else ""
            if not src:
                src = await img.first.get_attribute("data-src-pb") if await img.count() else ""

            sku = sku_from_src(src)

            products.append({
                "sku": sku,
                "quantity": 1,
                "status": status,
            })

        out.append({
            "adress_pvz": adress,
            "products": products,
        })

    return out


# ---------------- SAVE (МАССОВО) ----------------

def save_qr_png_near_excel(excel_path: str, *, phone10: str, code: str, qr_base64: str) -> str:
    xlsx = Path(excel_path).resolve()
    out_dir = xlsx.parent / "qr_images"
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_code = (code or "nocode").replace(" ", "")
    filename = f"{phone10}_{safe_code}.png"
    png_path = out_dir / filename

    if qr_base64:
        png_path.write_bytes(base64.b64decode(qr_base64))

    return str(png_path.relative_to(xlsx.parent)).replace("\\", "/")


def save_found_data_to_excel(
    xlsx_path: str,
    *,
    phone10: str,
    account_name: str,
    pvz_list: List[Dict[str, Any]],
    code: str,
    qr_base64: str,
) -> str:
    path = Path(xlsx_path)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "phone10",
            "account_name",
            "adress_pvz",
            "quantity",
            "sku_product",
            "status",
            "code",
            "qr",
        ])
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 9
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 24
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 10

    qr_rel = ""
    if qr_base64:
        qr_rel = save_qr_png_near_excel(str(path), phone10=phone10, code=code, qr_base64=qr_base64)

    for pvz in pvz_list:
        adress = pvz.get("adress_pvz", "")
        for p in pvz.get("products", []):
            ws.append([
                phone10,
                account_name,
                adress,
                int(p.get("quantity", 1)),
                p.get("sku", ""),
                p.get("status", ""),
                code,
                "",
            ])
            row = ws.max_row

            if qr_rel:
                cell = ws[f"H{row}"]
                cell.value = "QR.png"
                cell.hyperlink = qr_rel
                cell.font = Font(color="0000FF", underline="single")

    wb.save(path)
    return str(path)


# ✅ НОВОЕ: сохранить результаты ПОСЛЕ того как собрал все аккаунты
def save_mass_results_to_excel(xlsx_path: str, all_results: List[Dict[str, Any]]) -> str:
    """
    all_results — это список словарей, которые возвращает QRcodeScenario.run()
    """
    # (опционально) чтобы каждый массовый запуск начинал новый файл:
    # Path(xlsx_path).unlink(missing_ok=True)

    for r in all_results:
        save_found_data_to_excel(
            xlsx_path,
            phone10=r.get("phone10", ""),
            account_name=r.get("account_name", ""),
            pvz_list=r.get("pvz_list", []) or [],
            code=r.get("code", ""),
            qr_base64=r.get("qr_base64", ""),
        )
    return str(Path(xlsx_path).resolve())



