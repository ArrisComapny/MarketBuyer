import base64
import zipfile
import shutil
from datetime import datetime

from pathlib import Path

from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook

from domain.dtos import QrResult


def save_qr_png_near_excel(excel_path: str, *, phone10: str, code: str, qr_base64: str) -> str:
    xlsx = Path(excel_path).resolve()
    out_dir = xlsx.parent / "qr_images"
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_code = (code or "nocode").replace(" ", "")
    filename = f"{phone10}_{safe_code}.png"
    png_path = out_dir / filename

    if qr_base64:
        png_path.write_bytes(base64.b64decode(qr_base64))

    return str(png_path.relative_to(xlsx.parent)).replace("\\", "/")

def save_found_data_to_excel(xlsx_path: str, data: QrResult) -> str:
    path = Path(xlsx_path)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "phone10",
            "account_name",
            "adress_pvz",
            "quantity",
            "sku_product",
            "status",
            "code",
            "qr",
        ])
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 9
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 24
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 10

    qr_rel = ""
    if data.qr_base64:
        qr_rel = save_qr_png_near_excel(str(path), phone10=data.phone10, code=data.code, qr_base64=data.qr_base64)

    for pvz in data.pvz_list:
        for p in pvz.products:
            ws.append([
                data.phone10,
                data.account_name,
                pvz.address_pvz,
                int(p.quantity or 1),
                p.sku,
                p.status,
                data.code,
                "",
            ])
            row = ws.max_row

            if qr_rel:
                cell = ws[f"H{row}"]
                cell.value = "QR.png"
                cell.hyperlink = qr_rel
                cell.font = Font(color="0000FF", underline="single")

    wb.save(path)
    return str(path)

def save_mass_results_to_excel(xlsx_path: str, all_results: list[QrResult]) -> str:
    for r in all_results:
        save_found_data_to_excel(xlsx_path, r)
    return str(Path(xlsx_path).resolve())

def make_zip(temp_dir: Path, export_root: Path) -> str:
    """
    Создаёт архив из содержимого temp_dir,
    сохраняет архив в export_root,
    после чего удаляет temp_dir.
    """

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    zip_path = export_root / f"qr_export_{timestamp}.zip"

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in temp_dir.rglob("*"):
            if f.is_file():
                zf.write(f, f.relative_to(temp_dir))

    # удаляем временную папку полностью
    shutil.rmtree(temp_dir, ignore_errors=True)

    return str(zip_path)