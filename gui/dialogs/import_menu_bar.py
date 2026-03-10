from __future__ import annotations

import asyncio

import core.app as app_core

from PySide6.QtCore import Qt
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from PySide6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QComboBox
from PySide6.QtWidgets import QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem, QHeaderView

from database.repositories import AccountRepo, UsersAccountsRepo, UserRepo

from utils.phone import phone_to_10_digits
from utils.random_tools import pick_name_gender, pick_user_agent


class ImportMenuBarDialog(QDialog):
    """Окно импорта Excel."""

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Импорт из Excel")
        self.resize(520, 820)
        self.setWindowModality(Qt.WindowModality.ApplicationModal)

        self._file_path: str | None = None

        parent_user = getattr(parent, "user", None)
        self.is_admin = bool(parent_user and getattr(parent_user, "role", "") == "admin")

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        top = QHBoxLayout()
        self.lbl_file = QLabel("Файл: не выбран")
        self.btn_pick = QPushButton("Выбрать Excel…")
        self.btn_pick.clicked.connect(self.pick_file)

        top.addWidget(self.lbl_file, 1)
        top.addWidget(self.btn_pick, 0)
        root.addLayout(top)

        # --- Менеджер (только для админа) ---
        self.manager_combo = None
        if self.is_admin:
            manager_row = QHBoxLayout()
            self.lbl_manager = QLabel("Менеджер:")
            self.manager_combo = QComboBox()
            self.manager_combo.addItem("Выберите менеджера", None)

            manager_row.addWidget(self.lbl_manager, 0)
            manager_row.addWidget(self.manager_combo, 1)
            root.addLayout(manager_row)

            asyncio.create_task(self._load_managers())

        self.table = QTableWidget(self)
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Телефон", "Статус"])
        self.table.setRowCount(0)
        self.table.verticalHeader().setDefaultSectionSize(24)

        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)

        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(1, 140)

        root.addWidget(self.table, 1)

        bottom = QHBoxLayout()
        self.lbl_info = QLabel("0 строк")

        self.btn_import = QPushButton("Импортировать")
        self.btn_import.setMinimumHeight(40)
        self.btn_import.setEnabled(False)
        self.btn_import.clicked.connect(self.on_import_clicked)

        self.btn_close = QPushButton("Закрыть")
        self.btn_close.setMinimumHeight(40)
        self.btn_close.clicked.connect(self.reject)

        bottom.addWidget(self.lbl_info, 1)
        bottom.addWidget(self.btn_import, 0)
        bottom.addWidget(self.btn_close, 0)
        root.addLayout(bottom)



    def pick_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Выберите Excel файл", "", "Excel (*.xlsx)")
        if not path:
            return

        self._file_path = path
        self.lbl_file.setText(f"Файл: {path}")

        try:
            values = self.load_first_column(path)
        except Exception as e:
            self.lbl_info.setText(f"Ошибка чтения Excel: {e}")
            self.btn_import.setEnabled(False)
            return

        self.table.setRowCount(0)

        added = 0
        seen_phone10: set[str] = set()

        for v in values:
            raw = (str(v).strip() if v is not None else "")
            if not raw:
                continue

            phone10 = phone_to_10_digits(raw)
            status = "ok"

            if not phone10:
                status = "bad_phone"
            elif phone10 in seen_phone10:
                status = "dup_in_file"
            else:
                seen_phone10.add(phone10)

            row = self.table.rowCount()
            self.table.insertRow(row)

            it_phone10 = QTableWidgetItem(phone10 or "")
            it_status = QTableWidgetItem(status)

            it_phone10.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)

            self.table.setItem(row, 0, it_phone10)
            self.table.setItem(row, 1, it_status)

            added += 1

        self.lbl_info.setText(f"Загружено строк: {added}")
        self.btn_import.setEnabled(added > 0)

    @staticmethod
    def load_first_column(path: str) -> list[str]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        out: list[str] = []
        for (cell_val,) in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            if cell_val is None:
                continue
            s = str(cell_val).strip()
            if s:
                out.append(s)

        wb.close()
        return out

    def on_import_clicked(self) -> None:
        if not self._file_path:
            return

        if self.is_admin and self.manager_combo is not None and not self.manager_combo.currentData():
            self.lbl_info.setText("Выберите менеджера для импорта")
            self.manager_combo.setFocus()
            return

        asyncio.create_task(self.import_async())

    async def import_async(self) -> None:
        self.btn_import.setEnabled(False)
        self.btn_pick.setEnabled(False)

        parent_user = getattr(self.parent(), "user", None)
        login = getattr(parent_user, "login", None)

        if self.is_admin:
            assigned_login = self.manager_combo.currentData() if self.manager_combo else None
        else:
            assigned_login = login

        added_ok = 0
        skipped_bad = 0
        skipped_dupfile = 0
        skipped_exists = 0
        failed_other = 0

        try:
            async with app_core.db.get_session() as session:
                for r in range(self.table.rowCount()):
                    phone10 = (self.table.item(r, 0).text().strip() if self.table.item(r, 0) else "")
                    st_item = self.table.item(r, 1)

                    if not phone10:
                        if st_item:
                            st_item.setText("skip_bad_phone")
                        skipped_bad += 1
                        continue

                    if st_item and st_item.text().strip() == "dup_in_file":
                        skipped_dupfile += 1
                        continue

                    try:
                        async with session.begin_nested():
                            name, gender = await pick_name_gender(session)
                            ua = await pick_user_agent(session)

                            await AccountRepo.add_account(
                                session,
                                phone10=phone10,
                                name=name,
                                gender=gender,
                                user_agent=ua,
                                comment=""
                            )
                            await session.flush()

                            if assigned_login:
                                await UsersAccountsRepo.set_users_accounts(
                                    session,
                                    phone10=phone10,
                                    login=assigned_login
                                )

                        if st_item:
                            st_item.setText("imported")
                        added_ok += 1

                    except IntegrityError:
                        if st_item:
                            st_item.setText("exists")
                        skipped_exists += 1

                    except Exception as e:
                        if st_item:
                            st_item.setText(f"error: {e}")
                        failed_other += 1

                await session.commit()

            parent = self.parent()
            if parent and hasattr(parent, "load_accounts"):
                parent.load_accounts()

        finally:
            self.btn_pick.setEnabled(True)
            self.btn_import.setEnabled(True)

            self.lbl_info.setText(
                f"Добавлено: {added_ok}\n"
                f"Битых: {skipped_bad}\n"
                f"Дублей в файле: {skipped_dupfile}\n"
                f"Уже есть: {skipped_exists}\n"
                f"Ошибок: {failed_other}\n"
                f"Импорт завершён"
            )

    async def _load_managers(self) -> None:
        try:
            async with app_core.db.get_session() as session:
                users = await UserRepo.get_all_managers(session)

            if self.manager_combo is None:
                return

            self.manager_combo.clear()
            self.manager_combo.addItem("Выберите менеджера", None)

            for user in users:
                self.manager_combo.addItem(f"{user.login} ({user.name})", user.login)

        except Exception as e:
            self.lbl_info.setText(f"Ошибка загрузки менеджеров: {e}")
